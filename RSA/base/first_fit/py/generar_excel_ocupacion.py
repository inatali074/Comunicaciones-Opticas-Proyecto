import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os

def parse_speed(demand_id):
    """
    Extrae la velocidad (e.g. 100G, 200G, 300G, 400G) del ID de la demanda.
    Ejemplo: 'roadm Bragado-roadm 9 De Julio_100G_1' -> '100G'
    """
    if not demand_id or pd.isna(demand_id) or str(demand_id).strip() == "":
        return ""
    
    match = re.search(r'_(\d+G)_', str(demand_id))
    if match:
        return match.group(1)
    
    # Fallback si no está el guion bajo
    match_fallback = re.search(r'(\d+G)', str(demand_id))
    if match_fallback:
        return match_fallback.group(1)
        
    return ""

def generar_excel():
    base_dir = "/home/maximo/opticas/TpOpticas/RSA/base/first_fit"
    csv_path = os.path.join(base_dir, "ocupacion_base_firstfit_V2.csv")
    xlsx_path = os.path.join(base_dir, "ocupacion_visual_firstfit.xlsx")
    
    if not os.path.exists(csv_path):
        print(f"ERROR: No se encontró el archivo CSV de origen: {csv_path}")
        return
        
    print(f"Leyendo ocupación desde: {csv_path}")
    df = pd.read_csv(csv_path)
    
    # Cargar el mapeo de nombres largos a IDs cortos
    mapping_csv_path = os.path.join(base_dir, "mapping_lightpaths_firstfit.csv")
    mapping_dict = {}
    if os.path.exists(mapping_csv_path):
        print(f"Cargando mapeo de lightpaths desde: {mapping_csv_path}")
        df_map = pd.read_csv(mapping_csv_path)
        for _, r_map in df_map.iterrows():
            mapping_dict[str(r_map['Nombre_Lightpath']).strip()] = (str(r_map['ID_Demanda']), str(r_map['Velocidad']), int(r_map['Instancia']))
    
    # Crear Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ocupación de Slots"
    
    # Asegurar que las cuadrículas estén visibles
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos de encabezado principal (Gris oscuro/Azul oscuro)
    font_header_title = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header_title = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    
    font_header_sub = Font(name="Calibri", size=9, bold=False, color="FFFFFF")
    fill_header_freq = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Azul intermedio
    fill_header_wave = PatternFill(start_color="41719C", end_color="41719C", fill_type="solid") # Azul más claro
    
    # Estilos de celdas de datos
    font_data = Font(name="Calibri", size=10)
    font_bold_data = Font(name="Calibri", size=10, bold=True)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Bordes
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    header_border_side = Side(border_style="thin", color="FFFFFF")
    border_header = Border(left=header_border_side, right=header_border_side, top=header_border_side, bottom=header_border_side)
    
    # Paleta de colores para velocidades (Rellenos y Fuentes)
    colors = {
        "100G": {
            "fill": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),  # Azul pastel
            "font": Font(name="Calibri", size=9, bold=True, color="1F4E78")
        },
        "200G": {
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Verde pastel
            "font": Font(name="Calibri", size=9, bold=True, color="375623")
        },
        "300G": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Amarillo pastel
            "font": Font(name="Calibri", size=9, bold=True, color="7F6000")
        },
        "400G": {
            "fill": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),  # Rosa/Rojo pastel
            "font": Font(name="Calibri", size=9, bold=True, color="C00000")
        },
        "libre": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),  # Gris muy claro
            "font": Font(name="Calibri", size=9, color="7F7F7F")
        }
    }
    
    # ----------------------------------------------------
    # Escribir Fila 1: Encabezados de Columnas y Slots
    # ----------------------------------------------------
    ws.cell(row=1, column=1, value="Nodo Origen").font = font_header_title
    ws.cell(row=1, column=1).fill = fill_header_title
    ws.cell(row=1, column=1).alignment = align_center
    ws.cell(row=1, column=1).border = border_header
    
    ws.cell(row=1, column=2, value="Nodo Destino").font = font_header_title
    ws.cell(row=1, column=2).fill = fill_header_title
    ws.cell(row=1, column=2).alignment = align_center
    ws.cell(row=1, column=2).border = border_header
    
    for s in range(1, 305):
        col_idx = s + 2
        cell = ws.cell(row=1, column=col_idx, value=f"Slot {s}")
        cell.font = font_header_title
        cell.fill = fill_header_title
        cell.alignment = align_center
        cell.border = border_header
        
    # ----------------------------------------------------
    # Escribir Fila 2: Frecuencia Central [THz]
    # ----------------------------------------------------
    # Combinar A2 y B2 para la etiqueta de Frecuencia
    ws.merge_cells("A2:B2")
    ws.cell(row=2, column=1, value="Frecuencia [THz]").font = font_header_title
    ws.cell(row=2, column=1).fill = fill_header_freq
    ws.cell(row=2, column=1).alignment = align_center
    ws.cell(row=2, column=1).border = border_header
    ws.cell(row=2, column=2).fill = fill_header_freq
    ws.cell(row=2, column=2).border = border_header
    
    f_min = 191.35  # THz
    for s in range(1, 305):
        col_idx = s + 2
        f_c = f_min + (s - 0.5) * 0.0125  # Frecuencia central en THz
        cell = ws.cell(row=2, column=col_idx, value=round(f_c, 5))
        cell.font = font_header_sub
        cell.fill = fill_header_freq
        cell.alignment = align_center
        cell.border = border_header
        cell.number_format = "0.00000"
        
    # ----------------------------------------------------
    # Escribir Fila 3: Longitud de Onda [nm]
    # ----------------------------------------------------
    # Combinar A3 y B3 para la etiqueta de Longitud de Onda
    ws.merge_cells("A3:B3")
    ws.cell(row=3, column=1, value="Longitud de Onda [nm]").font = font_header_title
    ws.cell(row=3, column=1).fill = fill_header_wave
    ws.cell(row=3, column=1).alignment = align_center
    ws.cell(row=3, column=1).border = border_header
    ws.cell(row=3, column=2).fill = fill_header_wave
    ws.cell(row=3, column=2).border = border_header
    
    c = 299792.458  # Velocidad de la luz adaptada a THz y nm (c_m_s / 1e12 * 1e9)
    for s in range(1, 305):
        col_idx = s + 2
        f_c = f_min + (s - 0.5) * 0.0125
        lambda_val = c / f_c
        cell = ws.cell(row=3, column=col_idx, value=round(lambda_val, 2))
        cell.font = font_header_sub
        cell.fill = fill_header_wave
        cell.alignment = align_center
        cell.border = border_header
        cell.number_format = "0.00"

    # ----------------------------------------------------
    # Escribir Datos de Ocupación
    # ----------------------------------------------------
    print("Escribiendo enlaces y ocupación de slots...")
    for idx, row in df.iterrows():
        row_idx = idx + 4  # Los datos empiezan en la fila 4
        
        # Nodos de origen y destino (quitar prefijo 'roadm ' para mayor limpieza)
        origen = str(row["Nodo_Origen"]).replace("roadm ", "").strip()
        destino = str(row["Nodo_Destino"]).replace("roadm ", "").strip()
        
        # Celda Origen
        cell_orig = ws.cell(row=row_idx, column=1, value=origen)
        cell_orig.font = font_bold_data
        cell_orig.alignment = align_left
        cell_orig.border = border_all
        
        # Celda Destino
        cell_dest = ws.cell(row=row_idx, column=2, value=destino)
        cell_dest.font = font_bold_data
        cell_dest.alignment = align_left
        cell_dest.border = border_all
        
        # Slots 1 a 304
        for s in range(1, 305):
            col_idx = s + 2
            val_original = row[f"Slot_{s}"]
            
            # Limpiar valor
            val_clean = str(val_original).strip()
            if pd.isna(val_original) or val_clean == "" or val_clean == "nan":
                speed = ""
                cell_val = ""
            else:
                if val_clean in mapping_dict:
                    id_demanda, speed_val, instancia = mapping_dict[val_clean]
                    cell_val = f"{id_demanda}-{speed_val}_{instancia}"
                    speed = speed_val
                else:
                    speed = parse_speed(val_original)
                    cell_val = val_original
                
            cell_slot = ws.cell(row=row_idx, column=col_idx)
            cell_slot.alignment = align_center
            cell_slot.border = border_all
            
            if speed:
                cell_slot.value = cell_val
                cell_slot.fill = colors[speed]["fill"]
                cell_slot.font = colors[speed]["font"]
            else:
                cell_slot.value = ""
                cell_slot.fill = colors["libre"]["fill"]
                cell_slot.font = colors["libre"]["font"]
                
    # Ajustar ancho de columnas A y B
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    
    # Ajustar columnas de slots para que quepa Dk-VelocidadG_Instancia sin recortarse
    for s in range(1, 305):
        col_letter = openpyxl.utils.get_column_letter(s + 2)
        ws.column_dimensions[col_letter].width = 11.5
        
    # Inmovilizar paneles en C4 (Filas 1-3 fijas, Columnas A y B fijas)
    ws.freeze_panes = "C4"
    
    # ----------------------------------------------------
    # Agregar Pestaña "Lista de Demandas" (Diccionario)
    # ----------------------------------------------------
    dict_csv_path = os.path.join(base_dir, "lista_demandas_firstfit.csv")
    if os.path.exists(dict_csv_path):
        print(f"Cargando diccionario desde: {dict_csv_path}")
        df_dict = pd.read_csv(dict_csv_path)
        ws_dict = wb.create_sheet(title="Lista de Demandas")
        ws_dict.views.sheetView[0].showGridLines = True
        
        # Estilos del diccionario
        fill_dict_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_dict_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_dict_data = Font(name="Calibri", size=10)
        font_dict_bold = Font(name="Calibri", size=10, bold=True)
        
        # Escribir cabeceras del diccionario
        headers = ["ID_Demanda", "Origen", "Destino", "Cantidad de Enlaces", "Velocidad [Gbps]", "Ruta"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_dict.cell(row=1, column=col_idx, value=h)
            cell.font = font_dict_header
            cell.fill = fill_dict_header
            cell.alignment = align_center
            cell.border = border_all
            
        # Escribir filas del diccionario
        for row_idx, row_data in df_dict.iterrows():
            r = row_idx + 2
            ws_dict.row_dimensions[r].height = 18
            for col_idx, col_name in enumerate(headers, 1):
                val = row_data[col_name]
                cell = ws_dict.cell(row=r, column=col_idx, value=val)
                cell.border = border_all
                
                if col_name == "ID_Demanda":
                    cell.font = font_dict_bold
                    cell.alignment = align_center
                elif col_name in ["Cantidad de Enlaces", "Velocidad [Gbps]"]:
                    cell.font = font_dict_data
                    cell.alignment = align_center
                else:
                    cell.font = font_dict_data
                    cell.alignment = align_left
                    
        # Inmovilizar cabecera del diccionario
        ws_dict.freeze_panes = "A2"
        
        # Auto-ajustar columnas del diccionario
        for col in ws_dict.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_dict.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Guardar archivo
    print(f"Guardando Excel en: {xlsx_path}")
    wb.save(xlsx_path)
    print("¡Proceso completado con éxito!")

if __name__ == "__main__":
    generar_excel()
