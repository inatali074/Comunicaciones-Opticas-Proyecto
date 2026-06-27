import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import os

def parse_speed(demand_id):
    """
    Extrae la velocidad (e.g. 100G, 200G, 300G, 400G) del ID de la demanda.
    """
    if not demand_id or pd.isna(demand_id) or str(demand_id).strip() == "":
        return ""
    
    match = re.search(r'_(\d+G)_', str(demand_id))
    if match:
        return match.group(1)
        
    match_fallback = re.search(r'(\d+G)', str(demand_id))
    if match_fallback:
        return match_fallback.group(1)
        
    return ""

def generar_analisis_nodal():
    base_dir = "/home/maximo/opticas/TpOpticas/RSA/base/first_fit"
    csv_path = os.path.join(base_dir, "ocupacion_base_firstfit_V2.csv")
    xlsx_path = os.path.join(base_dir, "analisis_nodal_visual_firstfit.xlsx")
    
    if not os.path.exists(csv_path):
        print(f"ERROR: No se encontró el archivo CSV de origen: {csv_path}")
        return
        
    print(f"Leyendo ocupación desde: {csv_path}")
    df = pd.read_csv(csv_path)
    
    # Cargar el mapeo de nombres largos a IDs cortos
    mapping_csv_path = os.path.join(base_dir, "mapping_lightpaths_firstfit.csv")
    mapping_dict = {}
    if os.path.exists(mapping_csv_path):
        print(f"Cargando mapeo de lightpaths desde: {mapping_csv_path}")
        df_map = pd.read_csv(mapping_csv_path)
        for _, r_map in df_map.iterrows():
            mapping_dict[str(r_map['Nombre_Lightpath']).strip()] = (str(r_map['ID_Demanda']), str(r_map['Velocidad']), int(r_map['Instancia']))
    
    # Nodos a analizar (Nombre legible y ID de ROADM en la topología)
    NODOS_INTERES = [
        ("BENAVÍDEZ", "roadm Benavidez"),
        ("CAMPANA", "roadm Campana"),
        ("ZÁRATE", "roadm Zarate"),
        ("CONCEPCIÓN DEL URUGUAY", "roadm Concepcion del Uruguay"),
        ("ROSARIO", "roadm Rosario"),
        ("SAN FABIÁN", "roadm San Fabian"),
        ("CAÑADA DE GÓMEZ", "roadm Cañada De Gomez")
    ]
    
    # Crear Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabla de Canales"
    
    # Asegurar que las cuadrículas estén visibles
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos de fuentes y rellenos
    font_node_title = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    font_col_label = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10)
    font_bold_data = Font(name="Calibri", size=9, bold=True)
    
    fill_node_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    fill_meta_header = PatternFill(start_color="41719C", end_color="41719C", fill_type="solid") # Azul intermedio
    fill_link_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # Azul secundario
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Bordes
    thin_side_white = Side(border_style="thin", color="FFFFFF")
    border_header = Border(left=thin_side_white, right=thin_side_white, top=thin_side_white, bottom=thin_side_white)
    
    thin_side_grey = Side(border_style="thin", color="D9D9D9")
    border_data = Border(left=thin_side_grey, right=thin_side_grey, top=thin_side_grey, bottom=thin_side_grey)
    
    # Paleta de colores pastel
    colors = {
        "100G": {
            "fill": PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"),
            "font": Font(name="Calibri", size=9, bold=True, color="1F4E78")
        },
        "200G": {
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "font": Font(name="Calibri", size=9, bold=True, color="375623")
        },
        "300G": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "font": Font(name="Calibri", size=9, bold=True, color="7F6000")
        },
        "400G": {
            "fill": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
            "font": Font(name="Calibri", size=9, bold=True, color="C00000")
        },
        "libre": {
            "fill": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
            "font": Font(name="Calibri", size=9, color="7F7F7F")
        }
    }
    
    # ----------------------------------------------------
    # Generar Estructura de Columnas
    # ----------------------------------------------------
    # Para cada nodo de interés, identificamos sus enlaces salientes en la asignación
    column_offset = 1
    node_column_ranges = [] # Guardará tuplas de (nombre_nodo, start_col, end_col, outgoing_links)
    
    for label, uid in NODOS_INTERES:
        # Enlaces salientes activos en el CSV para este nodo
        outgoing_rows = df[df["Nodo_Origen"] == uid]
        # Ordenamos los destinos alfabéticamente para prolijidad
        destinos = sorted(outgoing_rows["Nodo_Destino"].tolist())
        
        # Cada bloque de nodo tiene:
        # 3 columnas de metadatos (Slot, Wavelength, Frequency) + len(destinos) columnas de enlaces
        start_col = column_offset
        end_col = start_col + 2 + len(destinos)
        
        node_column_ranges.append((label, uid, start_col, end_col, destinos))
        
        # El próximo bloque empieza después de la columna en blanco
        column_offset = end_col + 2
        
    # Altura de filas de cabecera
    ws.row_dimensions[2].height = 28 # Fila del nombre del nodo
    ws.row_dimensions[3].height = 24 # Fila de los títulos de columna
    
    # ----------------------------------------------------
    # Escribir Encabezados (Filas 2 y 3)
    # ----------------------------------------------------
    for label, uid, start_col, end_col, destinos in node_column_ranges:
        # Fila 2: Nombre del Nodo Combinado
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell_node = ws.cell(row=2, column=start_col, value=label)
        cell_node.font = font_node_title
        cell_node.fill = fill_node_header
        cell_node.alignment = align_center
        cell_node.border = border_header
        
        # Ponderar bordes en las celdas combinadas de la fila 2
        for col in range(start_col, end_col + 1):
            ws.cell(row=2, column=col).border = border_header
            ws.cell(row=2, column=col).fill = fill_node_header
            
        # Fila 3: Cabeceras de columnas
        # Metadatos
        c_slot = ws.cell(row=3, column=start_col, value="Slot")
        c_slot.font = font_col_label
        c_slot.fill = fill_meta_header
        c_slot.alignment = align_center
        c_slot.border = border_header
        
        c_wave = ws.cell(row=3, column=start_col+1, value="Longitud de onda [nm]")
        c_wave.font = font_col_label
        c_wave.fill = fill_meta_header
        c_wave.alignment = align_center
        c_wave.border = border_header
        
        c_freq = ws.cell(row=3, column=start_col+2, value="Frecuencia [THz]")
        c_freq.font = font_col_label
        c_freq.fill = fill_meta_header
        c_freq.alignment = align_center
        c_freq.border = border_header
        
        # Enlaces
        for idx, dest_uid in enumerate(destinos):
            col_idx = start_col + 3 + idx
            dest_name = dest_uid.replace("roadm ", "").strip()
            
            c_link = ws.cell(row=3, column=col_idx, value=f"A {dest_name}")
            c_link.font = font_col_label
            c_link.fill = fill_link_header
            c_link.alignment = align_center
            c_link.border = border_header
            
    # ----------------------------------------------------
    # Escribir Datos de Slots (Filas 4 a 307)
    # ----------------------------------------------------
    f_min = 191.35  # THz
    c = 299792.458  # Constante de conversión THz -> nm
    
    print("Escribiendo slots y buscando asignación espectral...")
    for s in range(1, 305):
        row_idx = s + 3 # Comienza en fila 4
        ws.row_dimensions[row_idx].height = 18
        
        # Valores de física del slot
        f_c = f_min + (s - 0.5) * 0.0125
        lambda_val = c / f_c
        
        for label, uid, start_col, end_col, destinos in node_column_ranges:
            # Metadatos del slot para este nodo
            cell_slot = ws.cell(row=row_idx, column=start_col, value=s)
            cell_slot.font = font_bold_data
            cell_slot.alignment = align_center
            cell_slot.border = border_data
            
            cell_wave = ws.cell(row=row_idx, column=start_col+1, value=round(lambda_val, 2))
            cell_wave.font = font_data
            cell_wave.alignment = align_center
            cell_wave.border = border_data
            cell_wave.number_format = "0.00"
            
            cell_freq = ws.cell(row=row_idx, column=start_col+2, value=round(f_c, 5))
            cell_freq.font = font_data
            cell_freq.alignment = align_center
            cell_freq.border = border_data
            cell_freq.number_format = "0.00000"
            
            # Asignaciones de canales/enlaces
            for idx, dest_uid in enumerate(destinos):
                col_idx = start_col + 3 + idx
                
                # Buscar fila en el dataframe original
                match_rows = df[(df["Nodo_Origen"] == uid) & (df["Nodo_Destino"] == dest_uid)]
                
                speed = ""
                cell_str_val = ""
                if not match_rows.empty:
                    val_original = match_rows.iloc[0][f"Slot_{s}"]
                    val_clean = str(val_original).strip()
                    if not pd.isna(val_original) and val_clean != "" and val_clean != "nan":
                        if val_clean in mapping_dict:
                            id_demanda, speed_val, instancia = mapping_dict[val_clean]
                            cell_str_val = f"{id_demanda}-{speed_val}_{instancia}"
                            speed = speed_val
                        else:
                            speed = parse_speed(val_original)
                            cell_str_val = val_original
                    
                cell_val = ws.cell(row=row_idx, column=col_idx)
                cell_val.alignment = align_center
                cell_val.border = border_data
                
                if speed:
                    cell_val.value = cell_str_val
                    cell_val.fill = colors[speed]["fill"]
                    cell_val.font = colors[speed]["font"]
                else:
                    cell_val.value = ""
                    cell_val.fill = colors["libre"]["fill"]
                    cell_val.font = colors["libre"]["font"]
                    
    # ----------------------------------------------------
    # Aplicar Bordes Estilo Caja a las Demandas
    # ----------------------------------------------------
    print("Aplicando bordes estilo caja a los bloques de demandas...")
    side_black_thick = Side(border_style="medium", color="000000")
    
    for label, uid, start_col, end_col, destinos in node_column_ranges:
        for idx, dest_uid in enumerate(destinos):
            col_idx = start_col + 3 + idx
            
            current_val = None
            block_start = None
            
            for r in range(4, 308):
                val = ws.cell(row=r, column=col_idx).value
                if val and str(val).strip() != "":
                    if val != current_val:
                        if current_val is not None:
                            for br in range(block_start, r):
                                cell = ws.cell(row=br, column=col_idx)
                                cell.border = Border(
                                    left=side_black_thick,
                                    right=side_black_thick,
                                    top=side_black_thick if br == block_start else None,
                                    bottom=side_black_thick if br == r - 1 else None
                                )
                        current_val = val
                        block_start = r
                else:
                    if current_val is not None:
                        for br in range(block_start, r):
                            cell = ws.cell(row=br, column=col_idx)
                            cell.border = Border(
                                left=side_black_thick,
                                right=side_black_thick,
                                top=side_black_thick if br == block_start else None,
                                bottom=side_black_thick if br == r - 1 else None
                            )
                        current_val = None
                        block_start = None
            
            if current_val is not None:
                for br in range(block_start, 308):
                    cell = ws.cell(row=br, column=col_idx)
                    cell.border = Border(
                        left=side_black_thick,
                        right=side_black_thick,
                        top=side_black_thick if br == block_start else None,
                        bottom=side_black_thick if br == 307 else None
                    )

    # ----------------------------------------------------
    # Formatear Ancho de Columnas
    # ----------------------------------------------------
    # Configurar anchos de columna fijos para legibilidad
    for label, uid, start_col, end_col, destinos in node_column_ranges:
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col)].width = 7    # Slot
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col+1)].width = 11  # Wavelength
        ws.column_dimensions[openpyxl.utils.get_column_letter(start_col+2)].width = 11  # Frequency
        
        # Enlaces
        for idx in range(len(destinos)):
            col_idx = start_col + 3 + idx
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 16 # Enlace
            
        # Columna vacía de separación posterior (si no es el último bloque)
        if end_col < column_offset - 2:
            sep_col = end_col + 1
            ws.column_dimensions[openpyxl.utils.get_column_letter(sep_col)].width = 3
            
    # Inmovilizar paneles en fila 4 (los encabezados siempre fijos)
    ws.freeze_panes = "A4"
    
    # ----------------------------------------------------
    # Agregar Pestaña "Lista de Demandas" (Diccionario)
    # ----------------------------------------------------
    dict_csv_path = os.path.join(base_dir, "lista_demandas_firstfit.csv")
    if os.path.exists(dict_csv_path):
        print(f"Cargando diccionario desde: {dict_csv_path}")
        df_dict = pd.read_csv(dict_csv_path)
        ws_dict = wb.create_sheet(title="Lista de Demandas")
        ws_dict.views.sheetView[0].showGridLines = True
        
        # Estilos del diccionario
        fill_dict_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_dict_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_dict_data = Font(name="Calibri", size=10)
        font_dict_bold = Font(name="Calibri", size=10, bold=True)
        
        # Escribir cabeceras del diccionario
        headers = ["ID_Demanda", "Origen", "Destino", "Cantidad de Enlaces", "Velocidad [Gbps]", "Ruta"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_dict.cell(row=1, column=col_idx, value=h)
            cell.font = font_dict_header
            cell.fill = fill_dict_header
            cell.alignment = align_center
            cell.border = border_data
            
        # Escribir filas del diccionario
        for row_idx, row_data in df_dict.iterrows():
            r = row_idx + 2
            ws_dict.row_dimensions[r].height = 18
            for col_idx, col_name in enumerate(headers, 1):
                val = row_data[col_name]
                cell = ws_dict.cell(row=r, column=col_idx, value=val)
                cell.border = border_data
                
                if col_name == "ID_Demanda":
                    cell.font = font_dict_bold
                    cell.alignment = align_center
                elif col_name in ["Cantidad de Enlaces", "Velocidad [Gbps]"]:
                    cell.font = font_dict_data
                    cell.alignment = align_center
                else:
                    cell.font = font_dict_data
                    cell.alignment = align_left
                    
        # Inmovilizar cabecera del diccionario
        ws_dict.freeze_panes = "A2"
        
        # Auto-ajustar columnas del diccionario
        for col in ws_dict.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_dict.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Guardar archivo
    print(f"Guardando Excel Nodal en: {xlsx_path}")
    wb.save(xlsx_path)
    print("¡Análisis nodal visual creado exitosamente!")

if __name__ == "__main__":
    generar_analisis_nodal()
